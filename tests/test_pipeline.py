"""End-to-end тест пайплайна (стадии 1–5 + xlsx) — полностью офлайн.

Все внешние границы — фейки: коллекторы читают сохранённые HTML/JSON-фикстуры
через инъектированные фетчеры, движок отвечает фиксированным JSON (нормализация
при `web_search=False`, скоринг при `web_search=True`), эмбеддер — фейк-моделью.
Сеть не трогается; на выходе — валидный `.xlsx`.
"""

from __future__ import annotations

import json
import logging
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from job_agent.config import Config
from job_agent.dedup import SeenStore
from job_agent.embeddings import Embedder
from job_agent.engines.base import Engine
from job_agent.models import RawPost
from job_agent.output.xlsx import COLUMNS
from job_agent.pipeline import build_collectors, run_backfill, run_nightly, run_pipeline
from job_agent.websearch.base import Searcher, SearchResult

FIXTURES = Path(__file__).parent / "fixtures"
FAR_PAST = datetime(2000, 1, 1, tzinfo=UTC)


class FakeModel:
    """Фейк-модель эмбеддингов: любой текст → один и тот же вектор (косинус ~1)."""

    def embed(self, texts):
        for _ in texts:
            yield [1.0, 0.0, 0.0]


def _embedder() -> Embedder:
    return Embedder(model=FakeModel())


def _score_json(*, track: str, overall: int) -> str:
    return json.dumps(
        {
            "track": track,
            "company_analysis": "scaleup, b2b",
            "company_confidence": "medium",
            "requirements": {"must": ["Python"], "nice": ["k8s"]},
            "matching": [],
            "scores": {
                "must": 80,
                "nice": 50,
                "seniority": 70,
                "context": 65,
                "overall": overall,
                "map_fit": 60,
            },
            "score_method": "среднее",
            "gaps": {"critical": [], "strategic": ["масштаб"], "cosmetic": []},
            "to_reach_100": [],
            "verdict": {
                "should_apply": True,
                "type": "precise_fit",
                "hr_screening_probability": "high",
                "final_stage_probability": "medium",
                "summary": "точное попадание",
            },
        }
    )


class StubEngine(Engine):
    """Движок-заглушка: нормализация (web_search=False) и скоринг (web_search=True).

    Нормализация выдаёт по одной вакансии на пост с уникальным заголовком (счётчик),
    чтобы дедуп их не схлопывал. Скоринг возвращает валидный `ScoreResult`-JSON.
    """

    def __init__(self, track_name: str) -> None:
        self.track_name = track_name
        self.normalize_calls = 0
        self.score_calls = 0

    def complete(self, prompt: str, *, web_search: bool = False) -> str:
        if web_search:
            self.score_calls += 1
            return _score_json(track=self.track_name, overall=80)
        self.normalize_calls += 1
        n = self.normalize_calls
        return json.dumps(
            [
                {
                    "title": f"Product Manager {n}",
                    "company": f"Acme {n}",
                    "link_or_contact": "@hr",
                    "salary": "300к",
                    "description": "продуктовая роль, владение метриками",
                }
            ]
        )


def _config(tmp_path: Path, *, single_track: bool) -> Config:
    resume = tmp_path / "resume.md"
    resume.write_text("Продакт-менеджер, владение метриками, mobile", encoding="utf-8")
    tracks = [
        {
            "id": "main",
            "name": "Основной",
            "resume_path": "resume.md",
        }
    ]
    if not single_track:
        resume2 = tmp_path / "resume2.md"
        resume2.write_text("AI Product, ML-метрики", encoding="utf-8")
        tracks.append({"id": "ai", "name": "AI", "resume_path": "resume2.md"})
    data = {
        "version": 1,
        "tracks": tracks,
        "scoring_engine": "cli",
        "cli_tool": "claude",
        "output_mode": "table",
        "tg_channels": [{"handle": "forhirejobs", "private": False}],
        "use_aggregators": True,
    }
    return Config.model_validate(data)


def _collectors(config: Config):
    tme_html = (FIXTURES / "tme_channel.html").read_text(encoding="utf-8")
    vseti_html = (FIXTURES / "vseti_jobs.html").read_text(encoding="utf-8")
    getmatch_json = (FIXTURES / "getmatch_vacancies.json").read_text(encoding="utf-8")
    return build_collectors(
        config,
        public_fetcher=lambda url: tme_html,
        vseti_fetcher=lambda url: vseti_html,
        getmatch_fetcher=lambda url: getmatch_json,
    )


def test_end_to_end_produces_valid_xlsx(tmp_path) -> None:
    config = _config(tmp_path, single_track=True)
    engine = StubEngine(config.tracks[0].name)
    out = tmp_path / "result.xlsx"

    result = run_pipeline(
        config,
        since=FAR_PAST,
        output_path=out,
        base_dir=tmp_path,
        engine=engine,
        embedder=_embedder(),
        collectors=_collectors(config),
        seen_store=SeenStore(":memory:"),
    )

    # Стадии сошлись: собрали посты, все уникальны и прошли фильтр, все сошлись в xlsx.
    assert result.collected > 0
    assert result.after_filter == result.collected
    assert result.written == result.after_filter
    assert engine.score_calls == result.written

    # Файл существует и валиден.
    assert result.output_path == out
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    headers = [ws.cell(row=1, column=c + 1).value for c in range(len(COLUMNS))]
    assert headers == list(COLUMNS)
    # строк данных ровно столько, сколько записано
    assert ws.max_row == result.written + 1


def test_single_track_hides_direction_column(tmp_path) -> None:
    config = _config(tmp_path, single_track=True)
    engine = StubEngine(config.tracks[0].name)
    out = tmp_path / "result.xlsx"
    run_pipeline(
        config,
        since=FAR_PAST,
        output_path=out,
        base_dir=tmp_path,
        engine=engine,
        embedder=_embedder(),
        collectors=_collectors(config),
        seen_store=SeenStore(":memory:"),
    )
    wb = load_workbook(out)
    ws = wb.active
    letter = ws.cell(row=1, column=COLUMNS.index("направление") + 1).column_letter
    assert ws.column_dimensions[letter].hidden is True


def test_second_run_yields_zero_new(tmp_path) -> None:
    """Дедуп держит водяной знак: повторный прогон на тех же данных → ноль новых."""
    config = _config(tmp_path, single_track=True)
    db = tmp_path / "seen.db"

    def _run() -> int:
        engine = StubEngine(config.tracks[0].name)
        with SeenStore(db) as store:
            res = run_pipeline(
                config,
                since=FAR_PAST,
                base_dir=tmp_path,
                engine=engine,
                embedder=_embedder(),
                collectors=_collectors(config),
                seen_store=store,
            )
        return res.written

    first = _run()
    second = _run()
    assert first > 0
    assert second == 0


def test_logs_stage_counters(tmp_path, caplog) -> None:
    config = _config(tmp_path, single_track=True)
    engine = StubEngine(config.tracks[0].name)
    with caplog.at_level(logging.INFO, logger="job_agent.pipeline"):
        run_pipeline(
            config,
            since=FAR_PAST,
            base_dir=tmp_path,
            engine=engine,
            embedder=_embedder(),
            collectors=_collectors(config),
            seen_store=SeenStore(":memory:"),
        )
    assert any("собрано" in r.message and "после фильтра" in r.message for r in caplog.records)


class DateRecordingCollector:
    """Стаб-коллектор: фильтрует посты по `since` и запоминает запрошенную границу."""

    def __init__(self, posts):
        self._posts = posts  # list[(date, RawPost)]
        self.since_seen: list[datetime] = []

    def fetch(self, since: datetime):
        self.since_seen.append(since)
        return [post for date, post in self._posts if date >= since]


def test_nightly_seeds_then_uses_watermark(tmp_path) -> None:
    """Первый nightly засевает по lookback; второй берёт водяной знак как `since`."""
    config = _config(tmp_path, single_track=True)
    db = tmp_path / "seen.db"
    t0 = datetime(2026, 6, 1, tzinfo=UTC)
    posts = [(t0, RawPost(raw_text="вакансия", source="stub", url="u1"))]

    # Первый прогон: метки нет → since = now - lookback, попадает в окно.
    first_collector = DateRecordingCollector(posts)
    with SeenStore(db) as store:
        first = run_nightly(
            config,
            now=datetime(2026, 6, 10, tzinfo=UTC),
            lookback_days=14,
            base_dir=tmp_path,
            engine=StubEngine(config.tracks[0].name),
            embedder=_embedder(),
            collectors=[first_collector],
            seen_store=store,
        )
        assert store.get_watermark() == datetime(2026, 6, 10, tzinfo=UTC)

    assert first.collected == 1
    assert first.written == 1
    # since на первом прогоне отсчитан от lookback, не от эпохи.
    assert first_collector.since_seen == [datetime(2026, 5, 27, tzinfo=UTC)]

    # Второй прогон: водяной знак (2026-06-10) → старый пост вне окна → ноль новых.
    second_collector = DateRecordingCollector(posts)
    with SeenStore(db) as store:
        second = run_nightly(
            config,
            now=datetime(2026, 6, 11, tzinfo=UTC),
            base_dir=tmp_path,
            engine=StubEngine(config.tracks[0].name),
            embedder=_embedder(),
            collectors=[second_collector],
            seen_store=store,
        )
    assert second_collector.since_seen == [datetime(2026, 6, 10, tzinfo=UTC)]
    assert second.collected == 0
    assert second.written == 0


def test_nightly_second_run_same_data_zero_new(tmp_path) -> None:
    """Даже если пост снова попадает в окно, seen-store держит ноль новых."""
    config = _config(tmp_path, single_track=True)
    db = tmp_path / "seen.db"
    # Пост датирован FAR_PAST? нет — коллектор игнорирует since, всегда отдаёт пост.

    class AlwaysCollector:
        def fetch(self, since: datetime):
            return [RawPost(raw_text="вакансия", source="stub", url="u1")]

    def _run(now: datetime) -> int:
        with SeenStore(db) as store:
            res = run_nightly(
                config,
                now=now,
                lookback_days=14,
                base_dir=tmp_path,
                engine=StubEngine(config.tracks[0].name),
                embedder=_embedder(),
                collectors=[AlwaysCollector()],
                seen_store=store,
            )
        return res.written

    first = _run(datetime(2026, 6, 10, tzinfo=UTC))
    second = _run(datetime(2026, 6, 11, tzinfo=UTC))
    assert first == 1
    assert second == 0


def test_run_backfill_collects_recent(tmp_path) -> None:
    """`run_backfill` считает `since` от текущего момента; стаб-коллектор игнорит дату."""

    class StubCollector:
        def fetch(self, since: datetime):
            assert since.tzinfo is not None  # since считается tz-aware
            return [RawPost(raw_text="вакансия", source="stub", url="u1")]

    config = _config(tmp_path, single_track=True)
    engine = StubEngine(config.tracks[0].name)
    result = run_backfill(
        config,
        days=7,
        base_dir=tmp_path,
        engine=engine,
        embedder=_embedder(),
        collectors=[StubCollector()],
        seen_store=SeenStore(":memory:"),
    )
    assert result.collected == 1
    assert result.written == 1


class EnrichStubEngine(Engine):
    """Движок-заглушка с маршрутизацией по типу промта (нормализация/скоринг/обогащение).

    Скоринг — единственный вызов с `web_search=True`. Остальные (`web_search=False`)
    различаются по заголовку отрендеренного промта: нормализация, сопроводительное,
    контакт-поиск.
    """

    def __init__(self, track_name: str, *, overall: int) -> None:
        self.track_name = track_name
        self.overall = overall
        self.normalize_calls = 0

    def complete(self, prompt: str, *, web_search: bool = False) -> str:
        if web_search:
            return _score_json(track=self.track_name, overall=self.overall)
        if "Промт сопроводительного" in prompt:
            return "Здравствуйте! Подходящий кандидат под вашу вакансию."
        if "Промт контакт-поиска" in prompt:
            return json.dumps(
                {
                    "target_roles": ["Hiring Manager"],
                    "queries_used": ["site:linkedin.com Acme"],
                    "candidates": [
                        {"name": "Иван Рекрутёров", "role": "HR", "confidence": "medium"}
                    ],
                    "fallback_paths": ["HR на сайте"],
                    "draft_message": "Здравствуйте, увидел вакансию...",
                }
            )
        self.normalize_calls += 1
        n = self.normalize_calls
        return json.dumps(
            [
                {
                    "title": f"Product Manager {n}",
                    "company": f"Acme {n}",
                    "link_or_contact": "@hr",
                    "salary": "300к",
                    "description": "продуктовая роль, владение метриками",
                }
            ]
        )


class FakeSearcher(Searcher):
    """Фейк web-поиска: фиксированная выдача без сети."""

    def search(self, query: str, *, max_results: int = 5) -> list[SearchResult]:
        return [SearchResult(title="Acme careers", url="https://acme.example/jobs")]


def _enrich_config(tmp_path: Path) -> Config:
    resume = tmp_path / "resume.md"
    resume.write_text("Продакт-менеджер, владение метриками, mobile", encoding="utf-8")
    cover = tmp_path / "cover.md"
    cover.write_text("Здравствуйте, меня зовут {{name}}.", encoding="utf-8")
    data = {
        "version": 1,
        "tracks": [
            {
                "id": "main",
                "name": "Основной",
                "resume_path": "resume.md",
                "cover_template_path": "cover.md",
            }
        ],
        "scoring_engine": "cli",
        "cli_tool": "claude",
        "output_mode": "table",
        "tg_channels": [{"handle": "forhirejobs", "private": False}],
        "use_aggregators": True,
        "cover_letter_threshold": 70,
        "enable_contacts": True,
    }
    return Config.model_validate(data)


def test_enrichment_populates_cover_and_contacts(tmp_path) -> None:
    """Финалисты выше порога получают сопроводительное; при enable_contacts — контакты."""
    config = _enrich_config(tmp_path)
    engine = EnrichStubEngine(config.tracks[0].name, overall=80)

    result = run_pipeline(
        config,
        since=FAR_PAST,
        base_dir=tmp_path,
        engine=engine,
        embedder=_embedder(),
        collectors=_collectors(config),
        seen_store=SeenStore(":memory:"),
        searcher=FakeSearcher(),
    )

    assert result.written > 0
    for res in result.results:
        assert res.cover_letter == "Здравствуйте! Подходящий кандидат под вашу вакансию."
        assert res.contacts is not None
        assert res.contacts.candidates[0].name == "Иван Рекрутёров"


def test_enrichment_below_threshold_skips_cover(tmp_path) -> None:
    """Ниже порога сопроводительное не готовится (cover_letter=None)."""
    config = _enrich_config(tmp_path)
    engine = EnrichStubEngine(config.tracks[0].name, overall=50)

    result = run_pipeline(
        config,
        since=FAR_PAST,
        base_dir=tmp_path,
        engine=engine,
        embedder=_embedder(),
        collectors=_collectors(config),
        seen_store=SeenStore(":memory:"),
        searcher=FakeSearcher(),
    )

    assert result.written > 0
    for res in result.results:
        assert res.cover_letter is None
        # контакты не зависят от порога сопроводительного
        assert res.contacts is not None


def test_contacts_disabled_yields_none(tmp_path) -> None:
    """Без enable_contacts контакт-ассист не запускается (и searcher не нужен)."""
    config = _enrich_config(tmp_path)
    config = config.model_copy(update={"enable_contacts": False})
    engine = EnrichStubEngine(config.tracks[0].name, overall=80)

    result = run_pipeline(
        config,
        since=FAR_PAST,
        base_dir=tmp_path,
        engine=engine,
        embedder=_embedder(),
        collectors=_collectors(config),
        seen_store=SeenStore(":memory:"),
    )

    assert result.written > 0
    for res in result.results:
        assert res.cover_letter is not None
        assert res.contacts is None
