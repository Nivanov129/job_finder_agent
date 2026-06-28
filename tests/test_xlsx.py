"""Тесты XLSX-выхода (стадия 7) — без сети, через openpyxl in-memory."""

from __future__ import annotations

from datetime import datetime

from job_agent.models import (
    ContactCandidate,
    ContactResult,
    EnrichedResult,
    Gaps,
    Requirements,
    ScoreResult,
    Scores,
    Vacancy,
    Verdict,
)
from job_agent.output.xlsx import (
    COLUMNS,
    TRACK_COLUMN_INDEX,
    build_workbook,
    write_xlsx,
)
from job_agent.presentation import badge_colors, to_argb


def _enriched(
    *,
    title: str,
    overall: int,
    track: str = "Бэкенд",
    map_fit: int = 50,
    cover_letter: str | None = None,
    contacts: ContactResult | None = None,
) -> EnrichedResult:
    vacancy = Vacancy(
        title=title,
        company="Acme",
        link_or_contact="@hr",
        salary="300к",
        description="desc",
        source="@jobs",
        url="https://t.me/jobs/1",
        date=datetime(2026, 6, 1, 12, 0),
    )
    score = ScoreResult(
        track=track,
        company_analysis="scaleup",
        company_confidence="medium",
        requirements=Requirements(must=["Python"], nice=["k8s"]),
        matching=[],
        scores=Scores(
            must=80,
            nice=50,
            seniority=70,
            context=65,
            overall=overall,
            map_fit=map_fit,
        ),
        score_method="среднее",
        gaps=Gaps(critical=["нет k8s"], strategic=["масштаб"], cosmetic=[]),
        to_reach_100=[],
        verdict=Verdict(
            should_apply=True,
            type="precise_fit",
            hr_screening_probability="high",
            final_stage_probability="medium",
            summary="точное попадание",
        ),
    )
    return EnrichedResult(
        vacancy=vacancy,
        score=score,
        cover_letter=cover_letter,
        contacts=contacts,
    )


def test_headers_match_format_spec() -> None:
    wb = build_workbook([_enriched(title="A", overall=85)])
    ws = wb.active
    headers = [ws.cell(row=1, column=c + 1).value for c in range(len(COLUMNS))]
    assert headers == list(COLUMNS)


def test_rows_sorted_by_overall_desc() -> None:
    wb = build_workbook(
        [
            _enriched(title="low", overall=60),
            _enriched(title="high", overall=90),
            _enriched(title="mid", overall=75),
        ]
    )
    ws = wb.active
    title_col = COLUMNS.index("должность") + 1
    titles = [ws.cell(row=r, column=title_col).value for r in (2, 3, 4)]
    assert titles == ["high", "mid", "low"]


def test_overall_cell_fill_matches_band() -> None:
    wb = build_workbook(
        [
            _enriched(title="green", overall=85),
            _enriched(title="amber", overall=75),
            _enriched(title="grey", overall=60),
        ]
    )
    ws = wb.active
    overall_col = COLUMNS.index("резюме %") + 1
    # строки отсортированы убыв.: 85, 75, 60
    for row, overall in ((2, 85), (3, 75), (4, 60)):
        cell = ws.cell(row=row, column=overall_col)
        expected = to_argb(badge_colors(overall).bg)
        assert cell.fill.start_color.rgb == expected
        assert cell.value == overall


def test_track_column_hidden_when_single_track() -> None:
    wb = build_workbook([_enriched(title="A", overall=85)], is_single_track=True)
    ws = wb.active
    letter = ws.cell(row=1, column=TRACK_COLUMN_INDEX).column_letter
    assert ws.column_dimensions[letter].hidden is True


def test_track_column_visible_with_multiple_tracks() -> None:
    wb = build_workbook([_enriched(title="A", overall=85)], is_single_track=False)
    ws = wb.active
    letter = ws.cell(row=1, column=TRACK_COLUMN_INDEX).column_letter
    assert ws.column_dimensions[letter].hidden is False
    # значение направления попало в строку
    assert ws.cell(row=2, column=TRACK_COLUMN_INDEX).value == "Бэкенд"


def test_cover_and_contacts_rendered() -> None:
    contacts = ContactResult(
        candidates=[ContactCandidate(name="Иван", role="CTO", link="t.me/ivan")],
        draft_message="Здравствуйте!",
    )
    wb = build_workbook(
        [
            _enriched(
                title="A",
                overall=85,
                cover_letter="Сопроводительное письмо",
                contacts=contacts,
            )
        ]
    )
    ws = wb.active
    cover_col = COLUMNS.index("сопроводительное") + 1
    contacts_col = COLUMNS.index("контакты") + 1
    assert ws.cell(row=2, column=cover_col).value == "Сопроводительное письмо"
    contacts_val = ws.cell(row=2, column=contacts_col).value
    assert "Иван" in contacts_val
    assert "Здравствуйте!" in contacts_val


def test_empty_cover_and_contacts_are_blank() -> None:
    wb = build_workbook([_enriched(title="A", overall=85)])
    ws = wb.active
    cover_col = COLUMNS.index("сопроводительное") + 1
    contacts_col = COLUMNS.index("контакты") + 1
    assert ws.cell(row=2, column=cover_col).value == ""
    assert ws.cell(row=2, column=contacts_col).value == ""


def test_write_xlsx_creates_file(tmp_path) -> None:
    from openpyxl import load_workbook

    out = tmp_path / "result.xlsx"
    returned = write_xlsx([_enriched(title="A", overall=85)], out)
    assert returned == out
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "дата"
    assert ws.cell(row=2, column=COLUMNS.index("должность") + 1).value == "A"
