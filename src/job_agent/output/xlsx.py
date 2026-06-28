"""XLSX-выход (стадия 7): финалисты скоринга → таблица `.xlsx`.

Колонки — по разделу «Форматы выхлопа» плана:
`дата · источник · должность · компания · ссылка/контакт · зарплата ·
резюме % · карта % · направление · вердикт · гэпы · сопроводительное · контакты`.

Строки сортируются по `overall` (резюме %) убыванию. Ячейка «резюме %»
заливается цветом бейджа из `presentation.py` (единый источник, не дублируем
палитру). Колонка «направление» скрывается при единственном треке
(`is_single_track`) — логика не ветвится, колонка просто прячется.
"""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from ..models import EnrichedResult
from ..presentation import (
    DEFAULT_AMBER_MIN,
    DEFAULT_GREEN_MIN,
    badge_colors,
    to_argb,
    verdict_style,
)

__all__ = ["COLUMNS", "TRACK_COLUMN_INDEX", "write_xlsx", "build_workbook"]

#: Заголовки колонок в порядке вывода (см. «Форматы выхлопа»).
COLUMNS: tuple[str, ...] = (
    "дата",
    "источник",
    "должность",
    "компания",
    "ссылка/контакт",
    "зарплата",
    "резюме %",
    "карта %",
    "направление",
    "вердикт",
    "гэпы",
    "сопроводительное",
    "контакты",
)

#: 1-based индекс колонки «направление» (скрывается при единственном треке).
TRACK_COLUMN_INDEX = COLUMNS.index("направление") + 1
#: 1-based индекс колонки «резюме %» (заливается цветом бейджа).
_OVERALL_COLUMN_INDEX = COLUMNS.index("резюме %") + 1


def _fmt_date(result: EnrichedResult) -> str:
    date = result.vacancy.date
    return date.strftime("%Y-%m-%d") if date is not None else ""


def _fmt_link(result: EnrichedResult) -> str:
    vacancy = result.vacancy
    return vacancy.link_or_contact or vacancy.url or ""


def _fmt_gaps(result: EnrichedResult) -> str:
    """Гэпы одной строкой: критические первыми, затем стратегические."""
    gaps = result.score.gaps
    lines: list[str] = []
    for label, items in (
        ("Критично", gaps.critical),
        ("Стратегически", gaps.strategic),
        ("Косметика", gaps.cosmetic),
    ):
        for item in items:
            lines.append(f"{label}: {item}")
    return "\n".join(lines)


def _fmt_contacts(result: EnrichedResult) -> str:
    contacts = result.contacts
    if contacts is None:
        return ""
    parts: list[str] = []
    for cand in contacts.candidates:
        bits = [cand.name]
        if cand.role:
            bits.append(cand.role)
        if cand.link:
            bits.append(cand.link)
        parts.append(" · ".join(bits))
    if contacts.draft_message:
        parts.append(f"Черновик: {contacts.draft_message}")
    return "\n".join(parts)


def _row_values(result: EnrichedResult) -> list[object]:
    vacancy = result.vacancy
    scores = result.score.scores
    return [
        _fmt_date(result),
        vacancy.source or "",
        vacancy.title,
        vacancy.company or "",
        _fmt_link(result),
        vacancy.salary or "",
        scores.overall,
        scores.map_fit,
        result.score.track,
        result.score.verdict.summary,
        _fmt_gaps(result),
        result.cover_letter or "",
        _fmt_contacts(result),
    ]


def _sorted_results(results: Iterable[EnrichedResult]) -> list[EnrichedResult]:
    return sorted(results, key=lambda r: r.score.scores.overall, reverse=True)


def build_workbook(
    results: Sequence[EnrichedResult],
    *,
    is_single_track: bool = False,
    green_min: int = DEFAULT_GREEN_MIN,
    amber_min: int = DEFAULT_AMBER_MIN,
) -> Workbook:
    """Построить книгу xlsx из финалистов (без записи на диск).

    Сорт по `overall` убыв.; «резюме %» заливается цветом бейджа; колонка
    «направление» скрыта при `is_single_track`.
    """
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Подборка"

    header_font = Font(bold=True)
    for col, title in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font

    for row, result in enumerate(_sorted_results(results), start=2):
        for col, value in enumerate(_row_values(result), start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        overall = result.score.scores.overall
        colors = badge_colors(overall, green_min=green_min, amber_min=amber_min)
        fill = to_argb(colors.bg)
        badge = ws.cell(row=row, column=_OVERALL_COLUMN_INDEX)
        badge.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        badge.font = Font(color=to_argb(colors.fg), bold=True)
        # тон вердикта по типу с фолбэком «на грани» при overall ниже зоны
        vstyle = verdict_style(
            result.score.verdict.type, overall=overall, amber_min=amber_min
        )
        ws.cell(row=row, column=COLUMNS.index("вердикт") + 1).font = Font(
            color=to_argb(vstyle.tone_hex)
        )

    if is_single_track:
        ws.column_dimensions[ws.cell(row=1, column=TRACK_COLUMN_INDEX).column_letter]\
            .hidden = True

    return wb


def write_xlsx(
    results: Sequence[EnrichedResult],
    path: str | Path,
    *,
    is_single_track: bool = False,
    green_min: int = DEFAULT_GREEN_MIN,
    amber_min: int = DEFAULT_AMBER_MIN,
) -> Path:
    """Записать `.xlsx` финалистов на диск, вернуть путь."""
    path = Path(path)
    wb = build_workbook(
        results,
        is_single_track=is_single_track,
        green_min=green_min,
        amber_min=amber_min,
    )
    wb.save(path)
    return path
